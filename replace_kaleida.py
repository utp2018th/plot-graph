# モジュールのインポート場所を変えるとエラーが起きることがあるので注意

def get_file_name():
    import os, tkinter, tkinter.filedialog, tkinter.messagebox

    root = tkinter.Tk()
    root.withdraw()
    # 選択候補を拡張子xlsxに絞る
    filetype = [("", "*.xlsx")]
    dirpath = os.path.abspath(os.path.dirname(".{}input{}".format(os.sep,os.sep)))
    tkinter.messagebox.showinfo('ファイル選択', '分析するエクセルファイルを選択してください')
    # 選択したファイルのパスを取得
    filepath = tkinter.filedialog.askopenfilename(filetypes = filetype, initialdir = dirpath)
    return filepath

def get_array_from_xlsx(path):
    import openpyxl as px

    wb = px.load_workbook(path)
    sheet = wb.active
    data = []
    for rows in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        array = []
        for cell in rows:
            array.append(cell.value)
        data.append(array)
    return data

class Plot_data:
    def __init__(self):
        self.a = None
        self.b = None
        self.v_max = None
        self.k_cat = None
        self.k_m = None
        self.k2 = None
        self.rss = None
        self.x_values = None
        self.y_values = None

    def calc(self):
        self.k_cat = self.v_max / (4.0 * 10**(-3))
        self.k2 = self.k_cat / self.k_m

def main():
    path = get_file_name()
    data = get_array_from_xlsx(path)

    import numpy as np
    import matplotlib.pyplot as plt
    import plot_master2 as pm2
    import sys
    import csv

    box = []
    array_s = []
    for idx,row in enumerate(data):
        array_s.append(row[1])
        del row[1]
        array = pm2.Data(row)
        if idx == 0:
            x = array
        else:
            box.append(array)

    # 吸光度のプロットに必要な変数
    title = 'Absorbance to Time'
    xlabel = 'Time(sec)'
    ylabel = 'Absorbance'
    output_path = 'day3/Absorbance'
    sig_dig = '3'
    args = [title,xlabel,ylabel,output_path,sig_dig]

    # 吸光度の変化をプロットする
    graph = pm2.Graph_master(args)
    graph.make_graph(x,box,input_mode="2")
    plt.savefig(graph.output_path)
    plt.close()

    # 反応初速度を求める
    v0 = []
    for i in box:
        if i.fit_values.size == 1:
            v0.append(i.fit_values)
        else:
            v0.append(i.fit_values[0])
    v0 = np.array(v0)
    e492 = 85000*10**(-6) # 単位をμLに揃える
    v0 /= e492

    # プロットのテンプレとなる関数
    def plot_and_save(x,y,args,input_mode=None,max_x=None):
        x = pm2.Data(x)
        y = [""] + list(y)
        ys = [pm2.Data(y)]
        graph = pm2.Graph_master(args)
        graph.make_graph(x,ys,input_mode,max_x)
        plt.savefig(graph.output_path)
        plt.close()
        data = Plot_data()
        data.x_values = x.values
        for y in ys:
            data.a = y.fit_values[0]
            data.b = y.fit_values[1]
            data.rss = y.rss
            data.y_values = y.values
        return data

    data_array = []

    # ６点でミカエリスメンテン
    # 変数の定義
    title = 'Michaelis Menten Plot (6 dots)'
    xlabel = 'Substrate concentration (μM)'
    ylabel = 'Reaction rate (μM/sec)'
    output_path = 'day3/MichaelisMenten6'
    args = [title,xlabel,ylabel,output_path,sig_dig]

    mm6 = plot_and_save(array_s,v0,args,input_mode="3",max_x=500)
    mm6.v_max = mm6.a
    mm6.k_m = mm6.b
    mm6.calc()
    data_array.append(mm6)

    # 4点でミカエリスメンテン
    # 変数の定義
    title = 'Michaelis Menten Plot (4 dots)'
    xlabel = 'Substrate concentration (μM)'
    ylabel = 'Reaction rate (μM/sec)'
    output_path = 'day3/MichaelisMenten4'
    args = [title,xlabel,ylabel,output_path,sig_dig]

    mm4 = plot_and_save(array_s[:5],v0[:4],args,input_mode="3",max_x=500)
    mm4.v_max = mm4.a
    mm4.k_m = mm4.b
    mm4.calc()
    data_array.append(mm4)

    # Lineweaver-Burk Plot
    # 下準備
    var_x = np.array(array_s[1:])
    var_x = 1 / var_x
    var_y = 1 / v0
    var_x = [array_s[0]] + list(var_x)

    # 変数の定義
    title = 'Lineweaver-Burk Plot'
    xlabel = '1 / Substrate concentration (/μM)'
    ylabel = '1 / Reaction rate (sec/M)'
    output_path = 'day3/Lineweaver-Burk'
    args = [title,xlabel,ylabel,output_path,sig_dig]

    lb = plot_and_save(var_x,var_y,args,input_mode="2")
    lb.v_max = 1 / lb.b
    lb.k_m = lb.a / lb.b
    lb.calc()
    data_array.append(lb)

    plt.figure(figsize=(9,6),dpi=128)
    mm2lb_x = np.linspace(0,1,100)
    mm2lb_y = mm6.b/mm6.a * mm2lb_x + 1/mm6.a
    plt.scatter(lb.x_values,lb.y_values)
    plt.plot(mm2lb_x,mm2lb_y,label='M-M\ny={:.3e}x+{:.3e}'.format(mm6.a,mm6.b))
    lb_line = lb.a * mm2lb_x + lb.b
    plt.plot(mm2lb_x,lb_line,label='L-B\ny={:.3e}x+{:.3e}'.format(lb.a,lb.b))
    plt.title('M-M6 and L-B')
    plt.xlabel('1 / Substrate concentration (/μM)')
    plt.ylabel('1 / Reaction rate (sec/M)')
    plt.legend()
    plt.grid()
    plt.savefig('output/day3/MM6-and-LB')

    # Eadie-Hofstee Plot
    # 下準備
    s = np.array(array_s[1:])
    var_y = v0
    var_x = v0 / s
    var_x = [array_s[0]] + list(var_x)

    # 変数の定義
    title = 'Eadie-Hofstee Plot'
    xlabel = 'Reaction rate / Substrate concentration (sec^-1)'
    ylabel = 'Reaction rate (M/sec)'
    output_path = 'day3/Eadie-Hofstee'
    args = [title,xlabel,ylabel,output_path,sig_dig]

    eh = plot_and_save(var_x,var_y,args,input_mode="2")
    eh.v_max = eh.b
    eh.k_m = - eh.a
    eh.calc()
    data_array.append(eh)

    def str_i(num):
        return '{:.3e}'.format(num)

    with open('output/day3/compare-k.csv',mode='w') as f:
        f.write(',M-M 6,M-M 4,L-B,E-H\n')
        k_cat = ",".join([str_i(i.k_cat) for i in data_array])
        f.write('k_cat,{}\n'.format(k_cat))
        k_m = ",".join([str_i(i.k_m) for i in data_array])
        f.write('k__m,{}\n'.format(k_m))
        k2 = ",".join([str_i(i.k2) for i in data_array])
        f.write('k2,{}\n'.format(k2))
        rss = ",".join([str_i(i.rss) for i in data_array])
        f.write('rss,{}\n'.format(rss))


if __name__ == "__main__":
    main()
